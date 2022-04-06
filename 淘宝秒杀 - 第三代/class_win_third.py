from tkinter import *
from tkinter import ttk
import tkinter.messagebox as messagebox
import requests
import re
from random import choice
from time import sleep
from mitmproxy.tools._main import mitmdump
import threading
import os
import NetDataBase
from datetime import date, timedelta, datetime
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.options import Options
import platform
import json
from selenium.webdriver.support.wait import WebDriverWait
import LocalDatabase
import pickle
import xlrd
import openpyxl

# 主窗
class APP(Tk):
    def __init__(self):
        super().__init__()
        self.title("农村淘宝抢购")
        self.geometry('500x300')
        # info that need to be show in the table
        self.beforeComp = []
        # info of the goods in cart
        self.cartGoodInfo = []
        # info after comparing
        self.afterComp = []
        # the goods need to buy in a mintue
        self.buyList = []
        self.cookie_path = os.path.abspath(os.getcwd()) + os.sep + 'COOKIE'
        cookieExist = self.checkCookieOkay()
        try:
            if cookieExist:
                self.refresh()
        except:
            messagebox.showinfo("提示","请点击登录并扫码")
        Excel().writeTimeAsZero(self.afterComp)
        self.setupUI()
        # Proxy()

    def checkCookieOkay(self):
        try:
            f = open(self.cookie_path, 'rb')
            NetDataBase.COOKIE = pickle.load(f)
            f.close()
            return True
        except:
            messagebox.showinfo("提示","请点击登录并扫码")
            return False

    def setupUI(self):
        # welcome notes
        Label(self, text='欢迎使用淘宝秒杀~', bg='bisque', font=('楷体', 20),width = 400, height=2).pack()
        # add goods
        addG = Frame(self)
        addG.pack(fill='x')
        Label(addG, text='', font=('楷体', 14),width = 5, height=3).pack(side=LEFT)
        Label(addG, text='', font=('楷体', 12),width = 2, height=2).pack(side=RIGHT)
        Label(addG, text='没有提示的时候不需要点击登录按钮，直接点击开始抢购即可', font=('楷体', 12),width = 50, height=2).pack(side=RIGHT)
        # start and exit
        Label(self, text='', font=('楷体', 12),width = 4, height=2).pack(side=LEFT)
        Button(self, text="确认好了，开始抢购", font=('楷体', 12), width=25, height=1, command=self.secKill).pack(side=LEFT)
        Label(self, text='', font=('楷体', 12),width = 2, height=2).pack(side=LEFT)
        Button(self, text="刷新", font=('楷体', 12), width=10, height=1, command=self.refresh).pack(side=LEFT)
        Label(self, text='', font=('楷体', 12),width = 2, height=2).pack(side=LEFT)
        Button(self, text="登录", font=('楷体', 12), width=10, height=1, command=self.log).pack(side=LEFT)
    ################################# LOGIN #####################################
    def log(self):
        self.cartGoodInfo = Buy().web_login()
        self.updateAfterDatabase()
        self.beforeComp.clear()

    ################################# REFRESH ########################################
    def refresh(self):
        self.clearAll()
        self.cartGoodInfo = Buy().direct_requests()
        self.readDataFromExcel()
        self.updateAfterDatabase()

    def clearAll(self):
        self.beforeComp = []
        self.afterComp = []
        self.buyList = []
        self.cartGoodInfo = []

    def readDataFromExcel(self):
        self.beforeComp = Excel().readTable()

    def updateAfterDatabase(self):
        for good in self.beforeComp:
            for data in self.cartGoodInfo:
                if good['name_short'] in str(data['name_full']):
                    self.afterComp.append({
                        'brand': data['brand'],
                        'name_short': good['name_short'],
                        'name_full': data['name_full'],
                        'ID': data['ID'],
                        'price_now': float(data['price_now'])/100,
                        'ideal_price': float(good['ideal_price']),
                        'time': good['time'],
                        'price_gap': float(data['price_now'])/100 - float(good['ideal_price']),
                        'buy_status': data['buy_status'],
                        'checked': data['checked'],
                    })
                    break
        print(self.afterComp)


    def secKill(self):
        startBuy = self.startBrowser()
        while(True):
            try:
                startBuy = self.goToCart(startBuy)
                print(11)
            except:
                print('跳转购物车失败')
            # 通过时间排序需要抢购物品的先后顺序
            try:
                self.afterComp = sorted(self.afterComp, key=lambda x: x['time'], reverse=False)
                time = self.find_first_time()
                if time == 0:
                    print('所有物品抢购完毕')
                print(12)
            except:
                print('找到首个时间失败')
            # 未到达指定时间，疯狂刷新现在价格
            print(Time().time_arrived(time))
            print(Time().time_gap)
            try:
                while not Time().time_arrived(time):
                    self.refreshPrice()
                    sleep(5)
                print(13)
            except:
                print("未能刷新价格")
                pass
            # 生成需要抢购的物品列表为buyList
            try:
                while(len(self.buyList) <= 0):
                    self.getBuyList()
                print(14)
            except:
                print('未能生成购买列表')
                pass
            # 选择要抢购的物品
            try:
                self.select_item(self.buyList, startBuy)
                sleep(2)
                print(15)
            except:
                print('未能选择商品')
            # 若价格合理则开始抢购, 抢购结束后关闭页面
            try:
                while not Time().time_arrived(time): pass
                print(16)
            except:
                print('等待出错')
            try:
                self.click_buy(startBuy, time)
                flag = True
                print(17)
            except:
                print("抢购失败")
                flag = False
            # 将被抢购的商品替换状态并且进行下一次抢购
            if flag is True:
                self.replaceBuyStatus()

    def startBrowser(self):
        startBuy = Login()
        # 1 初始化浏览器
        startBuy.init_browser()
        return startBuy
    
    def goToCart(self, startBuy):
        # 2 打开淘宝购物车
        cookie = startBuy.browser.get("https://cart.taobao.com/cart.htm?t=1609822441959")
        startBuy.browser.delete_all_cookies()
        f = open(self.cookie_path, 'rb')
        cookies = pickle.load(f)
        f.close()
        for cookie in cookies:
            startBuy.browser.add_cookie(cookie)
        startBuy.browser.get("https://cart.taobao.com/cart.htm?t=1609822441959")
        return startBuy.browser
        
    def find_first_time(self):
        for item in self.afterComp:
            if item['time'] != 0:
                return item['time']
        return 0

    def refreshPrice(self):
        Buy().direct_requests()
        self.beforeComp = self.afterComp
        self.updateAfterDatabase()
        self.beforeComp = []


    def getBuyList(self):
        self.afterComp = sorted(self.afterComp, key=lambda x: x['time'], reverse=False)
        for good in self.afterComp:
            if Time().time_one_less_than_time_two(good['time'], self.afterComp[0]['time']) and good['price_gap'] <= 0 and good['time'] != 0:
                self.buyList.append(good)

    def select_item(self, buyList, startBrowser):
        if len(buyList) == 0:
            return
        for item in buyList:
                if not item['checked']:
                    label = "".join(('//*[@id="J_Item_', item['ID'], '"]/ul/li[1]/div/div/div/label'))
                    target = startBrowser.find_element_by_xpath(label) 
                    startBrowser.execute_script("arguments[0].scrollIntoView();", target)
                    sleep(0.8)
                    startBrowser.find_element_by_xpath(label).click()
    
    def click_buy(self, startBrowser, time):
        startBrowser.find_element_by_id("J_Go").click()

        while True:
            try:
                startBrowser.find_element_by_xpath('//*[@id="ctTmypB2bFulfilmentSelectPC_ctTmypB2bFulfilmentSelectPC1"]/div/div/div/span/span[1]').click()
                break
            except:
                pass

        try:
            startBrowser.find_element_by_xpath('/html/body/div[4]/ul/li[2]/div').click()
        except:
            pass

        startBrowser.find_element_by_link_text('提交订单').click()
        
    def replaceBuyStatus(self):
        
        if len(self.buyList) == 0:
            return
        Excel().writeTimeAsZero(self.buyList)
        self.buyList.clear()
    
            

class Buy():
    def __init__(self):
        self.infoList = []
        self.url = 'https://cart.taobao.com/cart.htm?t=1609822441959'
        self.cookie_path = os.path.abspath(os.getcwd()) + os.sep + 'COOKIE'
    
    def checkCookieOkay(self):
        try:
            f = open(self.cookie_path, 'rb')
            NetDataBase.COOKIE = pickle.load(f)
            f.close()
        except:
            pass

    def web_login(self):
        log = Login()
        log.start()
        self.checkCookieOkay()
        html = self.getHTMLText()
        self.grabInfo(html)
        return(self.infoList)

    def direct_requests(self):
        self.checkCookieOkay()
        html = self.getHTMLText()
        self.grabInfo(html)
        return(self.infoList)

    def getHTMLText(self):
        for i in range(5):
            try:
                session = requests.Session()
                self.addCookie(session)
                head = {"user-agent": self.getUserAgent()}
                r = session.get(self.url, headers = head, timeout=30)
                r.raise_for_status
                r.encoding = r.apparent_encoding
                if not self.checkBeenScraped(r.text):
                    return r.text
                else:
                    print("被反爬，请更新cookie")
            except:
                sleep(5)
        print('ERROR in getHTMLText')

    def grabInfo(self, html):
        try:
            pl = re.findall(r'\"price\"\:\{\"actual\"\:.*?\"oriPromo', html)
            tl = re.findall(r'\"skuId\".*?\"\,\"toBuy\"',html)
            idl = re.findall(r'\}\,\"cartId\".*?\,\"createTime\"',html)
            priceList = []
            titleList = []
            idList = []
            checkedList = []
            for item in pl:
                price = item.split('"now":')[1]
                price = price.split(',"oriPromo')[0]
                priceList.append(price)
            for item in tl:
                title = item.split('"title":"')[1]
                title = title.split('","toBuy"')[0]
                titleList.append(title)
            for item in idl:
                idAndChecked = item.split('","')
                id = idAndChecked[0]
                checked = idAndChecked[1]
                id = id.split('"cartId":"')[1]
                if 'false' in str(checked):
                    checked = False
                elif 'true' in str(checked):
                    checked = True
                else:
                    checked = None
                idList.append(id)
                checkedList.append(checked)
            if len(priceList) != len(titleList) and len(titleList) != len(idList) and len(idList) != len(checkedList):
                print('Unbalanced List')
                return
            self.storeData(priceList, titleList, idList, checkedList)
        except:
            print('ERROR in grabInfo')
            return ""

    def getCookie(self, num):
        if num < len(NetDataBase.COOKIE): return NetDataBase.COOKIE[num]
        return None
    
    def getUserAgent(self):
        return choice(NetDataBase.USER_AGENTS)
    
    def storeData(self, priceList, titleList, idList, checkedList):
        for i in range(0, len(priceList)):
            tempDict = {"name_full": titleList[i],
                        "brand":  self.checkBrand(titleList[i]),
                        "ID": idList[i],
                        "price_now": priceList[i],
                        "checked": checkedList[i],
                        "buy_status": None,
                        }
            self.infoList.append(tempDict)
    
    def checkBrand(self, title):
        for goods in NetDataBase.BRAND:
            for goodName in goods['names']:
                if goodName in str(title):
                    return goods['outputName']
        return ""

    def checkBeenScraped(self, html):
        try:
            loginPage = re.findall(r'\"登录页面\"改进建议', html)
            if len(loginPage) != 0:
                return True
            else: False
        except:
            print('ERROR in checkBeenScraped')
            return False
    
    def addCookie(self,session):
        jar = requests.cookies.RequestsCookieJar()
        for cookie in NetDataBase.COOKIE:
            jar.set(cookie['name'], cookie['value'])
        session.cookies = jar
        

class Login():
    def __init__(self):
        self.browser = None
        self.loginURL = 'https://login.taobao.com/member/login.jhtml?redirectURL=http%3A%2F%2Fcart.taobao.com%2Fcart.htm'
        self.cookie_path = os.path.abspath(os.getcwd()) + os.sep + 'COOKIE'
    def start(self):
        # 1 初始化浏览器
        self.init_browser()
        # 2 打开淘宝登录页
        self.browser.get(self.loginURL)
        self.browser.implicitly_wait(1)
        self.browser.find_element_by_xpath('//*[@id="login"]/div[1]/i').click()
        while True:
            try:
                self.browser.find_element_by_id("J_SelectAllCbx1")
                break
            except:
                pass
        # get cookie
        cookie = self.browser.get_cookies()
        f = open(self.cookie_path, 'wb')
        f.write(pickle.dumps(cookie))
        f.close()
        # quit
        self.browser.close()
        return

    def init_browser(self):
        # 返回当前进程的工作目录
        self.downloadPath = os.getcwd()
        # 检索当前目录下的chromeDriver
        CHROME_DRIVER = os.path.abspath(os.path.dirname(os.getcwd())) + os.sep + 'chromedriver' + os.sep
        if platform.system() == 'Windows':
            CHROME_DRIVER = os.getcwd() + os.sep + 'chromedriver.exe'
        if platform.system() == 'Linux':
            CHROME_DRIVER = os.getcwd() + os.sep + 'chromedriver'
        """
        初始化selenium浏览器
        :return:
        """
        options = Options()
        # options.add_argument("--headless")
        # prefs = {"profile.managed_default_content_settings.images": 2}
        # 1是加载图片，2是不加载图片
        prefs = {"profile.managed_default_content_settings.images": 1, 'download.default_directory': self.downloadPath}
        options.add_experimental_option("prefs", prefs)
        # 设置为开发者模式，防止被各大网站识别出来使用了Selenium
        options.add_experimental_option('excludeSwitches', ['enable-automation']) 
        # 代理地址
        # options.add_argument('--proxy-server=http://127.0.0.1:9000')
        # 禁止显示导航栏
        options.add_argument('disable-infobars')
        # 初始化窗口最大
        options.add_argument('--start-maximized')
        # ？？
        options.add_argument('--no-sandbox')
        # # 保持窗口不要退出
        # options.add_experimental_option("detach", True)

        # 启动浏览器
        self.browser = webdriver.Chrome(executable_path=CHROME_DRIVER, options=options)
        self.browser.implicitly_wait(1)

'''
class Proxy():
    def __init__(self):
        thread = threading.Thread(target =  self.start)
        # thread.setDaemon(True)
        thread.start()

    def start(self):
        os.system("mitmdump -s HttpProxy.py -p 9000")
'''
class Time():
    def __init__(self):
        pass
    def get_time(self):
        self.now = datetime.now().strftime('%Y/%m/%d %H:%M')

    def time_arrived(self, time):
        self.time_gap(time)
        if self.gap <= 0 or self.gapr >= 0:
            return True
        return False
    
    def time_less_than_120s(self, time):
        self.time_gap(time)
        if self.gap <= 120:
            return True
        return False

    def time_one_less_than_time_two(self, time1, time2):
        time1 = datetime.strptime(time1, '%Y/%m/%d %H:%M')
        time2 = datetime.strptime(time2, '%Y/%m/%d %H:%M')
        self.gap = (time1 - time2).total_seconds()
        self.gapr = (time2 - time1).total_seconds()
        if self.gap <= 0 or self.gapr >= 0:
            return True
        return False

    def time_gap(self, time):
        time = datetime.strptime(time, '%Y/%m/%d %H:%M')
        self.gap = (time - datetime.now()).total_seconds()
        self.gapr = (datetime.now() - time).total_seconds()

def isElementExist(browser, xpath):
    try:
        browser.find_element_by_xpath(xpath)
        return True
    except:
        print('ERROR CANNOT FIND ELEMENT %s'%xpath)
        return False

class Excel():
    def __init__(self):
        self.excel_path = os.path.abspath(os.getcwd()) + os.sep + '抢购信息.xlsx'

    def readTable(self):
        data = xlrd.open_workbook(self.excel_path)
        table = data.sheet_by_name('1')
        rows = table.nrows
        cols = table.ncols
        info = list()
        for rowNum in range(3,rows):
            rowValue = table.row_values(rowNum)
            time = xlrd.xldate.xldate_as_datetime(rowValue[2],0)
            time = datetime.strftime(time,'%Y/%m/%d %H:%M')
            tempInfo = {
                'brand': '等待爬取数据',
                'name_short': rowValue[0],
                'name_full': '等待爬取数据',
                'ID': '等待爬取数据',
                'price_now': '等待爬取数据',
                'ideal_price': rowValue[1],
                'buy_status': '等待爬取数据',
                'time': time,
                'price_gap': '等待爬取数据',
            }
            info.append(tempInfo)
        return info

    def writeTimeAsZero(self, changeList):
        tempList = []
        for item in changeList:
            tempList.append(item['name_short'])
        workbook=openpyxl.load_workbook(self.excel_path)
        worksheet=workbook.worksheets[0]
        rows = worksheet.max_row
        for changeName in tempList:
            for rowNum in range(4, rows+1):
                if changeName in str(worksheet.cell(row=rowNum, column=1).value):
                    print(worksheet.cell(rowNum, 1).value)
                    worksheet.cell(row=rowNum,column=3).value = '0'
        
        workbook.save('抢购信息.xlsx')

app = APP()
app.mainloop()